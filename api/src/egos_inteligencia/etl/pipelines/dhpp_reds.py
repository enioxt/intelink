"""
DHPP REDS Pipeline — EGOS Inteligência / Intelink
Processa planilhas REDS (SISPOL/PM-MG) de homicídios e porte/posse de arma.

Fontes:
  - resultados (6).xlsx  → 885 homicídios, 4800 envolvidos (36 colunas)
  - resultados (7).xlsx  → 1207 porte/posse arma, 3617 envolvidos (36 colunas)
  - resultados (5).xlsx  → 882 homicídios (21 colunas, sem pessoa — fallback)
  - resultados (8).xlsx  → 1207 porte/posse arma (21 colunas, sem pessoa — fallback)

Estratégia de identidade (padrão MG/REDS):
  1. CPF válido (Mod-11)    → MERGE em p.cpf             (conf 0.95)
  2. RG SSP-MG              → MERGE em p.rg_mg            (conf 0.90)
  3. Nome + Mãe + Nascimento → MERGE em p.reds_person_key (conf 0.88)
  4. Nome + Nascimento       → POSSIBLE_SAME_AS           (conf 0.70)

Extração do campo histórico:
  - Placas de veículos (formato antigo ABC1234 e Mercosul ABC1D23)
  - Calibres de armas (.38, 9mm, .380, .40, .357, .44, .45, 12, .308, 5.56, 7.62)
  - IMEIs (15 dígitos após "IMEI")
  - Telefones (DD)NNNNN-NNNN

Sacred Code: 000.111.369.963.1618
"""
from __future__ import annotations

import hashlib
import logging
import os
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl

from ..compat.base import Pipeline
from ..compat.loader import Neo4jBatchLoader
from ..compat.transforms import (
    format_cpf,
    normalize_name,
    strip_document,
    validate_cpf,
)

logger = logging.getLogger(__name__)

# ── Regex patterns ─────────────────────────────────────────────────────────────

# Vehicle plates: old ABC1234 and Mercosul ABC1D23
PLATE_RE = re.compile(
    r'\b([A-Z]{3}[-\s]?[0-9][A-Z0-9][0-9]{2})\b',
    re.IGNORECASE
)
# Calibers
CALIBER_RE = re.compile(
    r'\b(\.38|9\s?[Mm][Mm]|\.380|\.40S?W?|\.357|\.44|\.45|'
    r'12\s?(?:GAUGE|BORE|GA)?|\.308|5\.56|7\.62|\.22|\.32|\.25)\b'
)
# IMEI
IMEI_RE = re.compile(r'IMEI[:\s]*(\d{15})', re.IGNORECASE)
# Phone BR
PHONE_RE = re.compile(r'\((\d{2})\)\s*(\d{4,5}[-\s]?\d{4})')

# REDS number format: YYYY-NNNNNNNNN-NNN
REDS_RE = re.compile(r'(\d{4}-\d{9,12}-\d{3})')

# ── Column maps ────────────────────────────────────────────────────────────────

COLS_21 = [
    "numero_ocorrencia", "tipo_relatorio", "natureza_principal", "consumado",
    "data_hora_fato", "ano_fato", "mes_fato", "modo_acao", "meio_utilizado",
    "causa_presumida", "local_imediato", "cod_municipio", "municipio_fato",
    "bairro_fato", "localizacao_ocorrencia", "endereco_ocorrencia",
    "latitude", "longitude", "unidade_registro", "orgao_registro", "historico",
]

COLS_36 = [
    "numero_ocorrencia", "tipo_relatorio", "natureza_principal", "consumado",
    "modo_acao", "meio_utilizado", "causa_presumida", "local_imediato",
    "data_hora_fato", "ano_fato", "mes_fato", "cod_municipio", "municipio_fato",
    "bairro_fato", "localizacao_ocorrencia", "endereco_ocorrencia",
    "latitude", "longitude",
    # Person fields (18-35)
    "chave_envolvido", "tipo_envolvimento", "natureza_delito", "prisao_apreensao",
    "nome_envolvido", "sexo", "rg", "data_nascimento", "cpf",
    "nome_mae_envolvido", "grau_lesao", "municipio_envolvido", "bairro_envolvido",
    "endereco_envolvido", "email", "telefone_residencial", "telefone_comercial",
    "historico",
]

# Crime type normalization
CRIME_TYPE_MAP: dict[str, str] = {
    "HOMICIDIO": "HOMICIDIO",
    "PORTE ILEGAL ARMA DE FOGO/ACESSÓRIO/MUNIÇÃO DE USO PERMITIDO": "PORTE_ARMA",
    "POSSE ILEGAL ARMA DE FOGO/ACESSÓRIO/MUNIÇÃO DE USO PERMITIDO": "POSSE_ARMA",
    "POSSE/PORTE ILEGAL ARMA FOGO/MUNIC/ACESSO USO PROIB/RESTRITO": "PORTE_POSSE_RESTRITO",
    "PORTE ILEGAL DE ARMA DE FOGO DE USO PERMITIDO": "PORTE_ARMA",
    "POSSE/PORTE ILEGAL DE ARMADEFOGO/USO RESTRITO": "PORTE_POSSE_RESTRITO",
}

INVOLVEMENT_MAP: dict[str, str] = {
    "AUTOR": "AUTOR",
    "CO-AUTOR": "CO_AUTOR",
    "SUSPEITO": "SUSPEITO",
    "VITIMA DE ACAO CRIMINAL / CIVEL": "VITIMA",
    "VITIMA - OUTROS": "VITIMA",
    "VITIMA": "VITIMA",
    "TESTEMUNHA QUE PRESENCIOU OS FATOS": "TESTEMUNHA",
    "TESTEMUNHA DA AÇÃO DOS POLICIAIS/BOMBEIROS": "TESTEMUNHA",
    "TESTEMUNHA QUE TOMOU CONHECIMENTO": "TESTEMUNHA",
    "TESTEMUNHA DE APRESENTACAO": "TESTEMUNHA",
    "TESTEMUNHA QUE PRESENCIOU ACORDO": "TESTEMUNHA",
    "REPRESENTANTE LEGAL": "REPRESENTANTE_LEGAL",
    "SOLICITANTE": "SOLICITANTE",
    "SOCORRISTA": "OUTROS",
    "CONDUTOR DO VEICULO": "OUTROS",
    "CONDUTOR DE VEICULO E VITIMA": "VITIMA",
    "CUSTODIANTE": "OUTROS",
    "IGNORADO": "OUTROS",
    "OUTROS ": "OUTROS",
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def normalize_rg(rg: str | None) -> str | None:
    """Extract only digits from RG, stripping leading zeros. Returns None if empty."""
    if not rg:
        return None
    digits = re.sub(r"[^0-9]", "", str(rg)).lstrip("0")
    return digits if digits else None


def make_rg_mg(rg: str | None) -> str | None:
    """Create rg_mg key: '{digits}_MG'. None if no digits."""
    digits = normalize_rg(rg)
    return f"{digits}_MG" if digits else None


def make_reds_person_key(
    nome: str | None,
    nome_mae: str | None,
    data_nasc: str | None,
) -> str | None:
    """Stable hash key from nome + nome_mãe + data_nascimento.
    Returns None if we don't have enough signals.
    """
    nome_n = normalize_name(nome)
    mae_n = normalize_name(nome_mae)
    nasc = (data_nasc or "").strip()
    if not nome_n:
        return None
    raw = f"{nome_n}|{mae_n}|{nasc}"
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def normalize_date(val: Any) -> str | None:
    """Normalize date to DD/MM/YYYY string."""
    if not val:
        return None
    s = str(val).strip()
    # Already DD/MM/YYYY
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        return s
    # DD/MM/YYYY HH:MM:SS
    m = re.match(r"^(\d{2}/\d{2}/\d{4})", s)
    if m:
        return m.group(1)
    # YYYY-MM-DD
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return s


def extract_plates(text: str | None) -> list[str]:
    if not text:
        return []
    found = PLATE_RE.findall(text.upper())
    return list({p.replace(" ", "").replace("-", "").upper() for p in found})


def extract_calibers(text: str | None) -> list[str]:
    if not text:
        return []
    found = CALIBER_RE.findall(text)
    return list({c.upper().strip() for c in found})


def extract_imeis(text: str | None) -> list[str]:
    if not text:
        return []
    return list(set(IMEI_RE.findall(text)))


def normalize_crime_type(raw: str | None) -> str:
    if not raw:
        return "OUTRO"
    key = (raw or "").strip().upper()
    return CRIME_TYPE_MAP.get(key, key[:30])


def normalize_involvement(raw: str | None) -> str:
    if not raw:
        return "OUTROS"
    key = (raw or "").strip().upper()
    # Normalize trailing spaces
    return INVOLVEMENT_MAP.get(key, INVOLVEMENT_MAP.get(key.rstrip(), "OUTROS"))


@dataclass
class REDSOccurrence:
    reds_number: str
    type: str
    consumado: bool
    data_fato: str | None
    ano_fato: int | None
    mes_fato: str | None
    municipio: str
    bairro: str | None
    endereco: str | None
    latitude: float | None
    longitude: float | None
    modo_acao: str | None
    meio_utilizado: str | None
    causa_presumida: str | None
    local_imediato: str | None
    historico: str | None
    unidade_registro: str | None
    orgao_registro: str | None
    plates: list[str] = field(default_factory=list)
    calibers: list[str] = field(default_factory=list)
    imeis: list[str] = field(default_factory=list)
    source: str = "reds_patos_minas"


@dataclass
class REDSPerson:
    chave_envolvido: str            # Unique key per person-per-occurrence
    reds_number: str
    tipo_envolvimento: str          # Normalized
    nome: str | None
    nome_normalized: str | None
    sexo: str | None
    rg: str | None
    rg_mg: str | None
    cpf: str | None                 # Formatted if valid
    cpf_valid: bool
    nome_mae: str | None
    nome_mae_normalized: str | None
    data_nascimento: str | None
    bairro: str | None
    municipio: str | None
    endereco: str | None
    telefone: str | None
    grau_lesao: str | None
    prisao: str | None
    natureza_delito: str | None
    reds_person_key: str | None
    source: str


class DHPPRedsPipeline(Pipeline):
    """
    Processa planilhas REDS de Patos de Minas → Neo4j.

    Cria:
      :Occurrence nodes (cada BO/REDS)
      :Person nodes (cada envolvido, com dedup MG)
      :Location nodes (bairro + município)
      ENVOLVIDO_EM relationships
      OCORREU_EM relationships
    """

    name = "dhpp_reds"
    source_id = "reds_patos_minas"

    # Files in data_dir/dhpp/
    FILES_36COL = [
        ("resultados (6).xlsx", "reds_homicidio"),
        ("resultados (7).xlsx", "reds_arma_fogo"),
    ]
    FILES_21COL = [
        ("resultados (5).xlsx", "reds_homicidio"),
        ("resultados (8).xlsx", "reds_arma_fogo"),
    ]
    SUBDIR = "dhpp"

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        self.occurrences: dict[str, REDSOccurrence] = {}   # reds_number → occurrence
        self.persons: dict[str, REDSPerson] = {}            # chave_envolvido → person
        self.merge_candidates: list[dict] = []             # For POSSIBLE_SAME_AS review

    # ── Extract ────────────────────────────────────────────────────────────────

    def extract(self) -> None:
        data_path = Path(self.data_dir) / self.SUBDIR
        if not data_path.exists():
            raise FileNotFoundError(
                f"Data dir not found: {data_path}. "
                "Copy REDS xlsx files to data/dhpp/ and retry."
            )

        # Process 36-column files first (have person data)
        for filename, source in self.FILES_36COL:
            fp = data_path / filename
            if fp.exists():
                logger.info("[dhpp_reds] Reading %s (36-col)...", filename)
                self._read_36col(fp, source)
            else:
                logger.warning("[dhpp_reds] File not found (skip): %s", fp)

        # Process 21-column files (occurrences without person data)
        # Only add occurrences not already seen in 36-col files
        for filename, source in self.FILES_21COL:
            fp = data_path / filename
            if fp.exists():
                logger.info("[dhpp_reds] Reading %s (21-col)...", filename)
                self._read_21col(fp, source)
            else:
                logger.warning("[dhpp_reds] File not found (skip): %s", fp)

        logger.info(
            "[dhpp_reds] Extracted: %d occurrences, %d envolvidos",
            len(self.occurrences),
            len(self.persons),
        )

    def _read_36col(self, fp: Path, source: str) -> None:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb["Resultados"]
        rows_read = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if self.limit and rows_read >= self.limit:
                break
            d = dict(zip(COLS_36, row))
            reds = str(d.get("numero_ocorrencia") or "").strip()
            if not reds:
                continue

            # Upsert occurrence
            if reds not in self.occurrences:
                self.occurrences[reds] = self._build_occurrence(d, source)

            # Build person if we have name
            chave = str(d.get("chave_envolvido") or "").strip()
            nome = (d.get("nome_envolvido") or "").strip()
            if chave and nome:
                self.persons[chave] = self._build_person(d, reds, source)

            rows_read += 1
        wb.close()
        logger.info("[dhpp_reds] %s: %d rows read", fp.name, rows_read)

    def _read_21col(self, fp: Path, source: str) -> None:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb["Resultados"]
        rows_read = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if self.limit and rows_read >= self.limit:
                break
            d = dict(zip(COLS_21, row))
            reds = str(d.get("numero_ocorrencia") or "").strip()
            if not reds or reds in self.occurrences:
                rows_read += 1
                continue  # Skip if already loaded from 36-col
            self.occurrences[reds] = self._build_occurrence(d, source)
            rows_read += 1
        wb.close()
        logger.info("[dhpp_reds] %s: %d rows read", fp.name, rows_read)

    def _build_occurrence(self, d: dict, source: str) -> REDSOccurrence:
        historico = (d.get("historico") or "")
        ano = d.get("ano_fato")
        return REDSOccurrence(
            reds_number=str(d.get("numero_ocorrencia") or "").strip(),
            type=normalize_crime_type(d.get("natureza_principal")),
            consumado=(str(d.get("consumado") or "N").strip().upper() == "S"),
            data_fato=normalize_date(d.get("data_hora_fato")),
            ano_fato=int(ano) if ano else None,
            mes_fato=str(d.get("mes_fato") or "").strip() or None,
            municipio=normalize_name(d.get("municipio_fato") or "PATOS DE MINAS"),
            bairro=normalize_name(d.get("bairro_fato")),
            endereco=(d.get("endereco_ocorrencia") or "").strip() or None,
            latitude=_safe_float(d.get("latitude")),
            longitude=_safe_float(d.get("longitude")),
            modo_acao=(d.get("modo_acao") or "").strip()[:500] or None,
            meio_utilizado=(d.get("meio_utilizado") or "").strip() or None,
            causa_presumida=(d.get("causa_presumida") or "").strip() or None,
            local_imediato=(d.get("local_imediato") or "").strip() or None,
            historico=historico[:5000] if historico else None,
            unidade_registro=(d.get("unidade_registro") or "").strip() or None,
            orgao_registro=(d.get("orgao_registro") or "").strip() or None,
            plates=extract_plates(historico),
            calibers=extract_calibers(historico),
            imeis=extract_imeis(historico),
            source=source,
        )

    def _build_person(self, d: dict, reds_number: str, source: str) -> REDSPerson:
        nome = (d.get("nome_envolvido") or "").strip()
        nome_normalized = normalize_name(nome)
        cpf_raw = str(d.get("cpf") or "").strip()
        cpf_digits = strip_document(cpf_raw)
        cpf_valid = validate_cpf(cpf_digits)
        cpf_fmt = format_cpf(cpf_digits) if cpf_valid else None

        rg_raw = str(d.get("rg") or "").strip()
        rg_mg = make_rg_mg(rg_raw)

        nome_mae = (d.get("nome_mae_envolvido") or "").strip() or None
        nome_mae_n = normalize_name(nome_mae)

        nasc = normalize_date(d.get("data_nascimento"))
        reds_key = make_reds_person_key(nome_normalized, nome_mae_n, nasc)

        tipo = normalize_involvement(d.get("tipo_envolvimento"))
        prisao_raw = (d.get("prisao_apreensao") or "").strip().upper()
        prisao = _normalize_prisao(prisao_raw)

        tel = _best_phone(d.get("telefone_residencial"), d.get("telefone_comercial"))

        return REDSPerson(
            chave_envolvido=str(d.get("chave_envolvido") or "").strip(),
            reds_number=reds_number,
            tipo_envolvimento=tipo,
            nome=nome or None,
            nome_normalized=nome_normalized or None,
            sexo=(str(d.get("sexo") or "").strip().upper() or None),
            rg=normalize_rg(rg_raw),
            rg_mg=rg_mg,
            cpf=cpf_fmt,
            cpf_valid=cpf_valid,
            nome_mae=nome_mae,
            nome_mae_normalized=nome_mae_n or None,
            data_nascimento=nasc,
            bairro=normalize_name(d.get("bairro_envolvido")),
            municipio=normalize_name(d.get("municipio_envolvido")),
            endereco=(d.get("endereco_envolvido") or "").strip() or None,
            telefone=tel,
            grau_lesao=(d.get("grau_lesao") or "").strip() or None,
            prisao=prisao,
            natureza_delito=(d.get("natureza_delito") or "").strip()[:200] or None,
            reds_person_key=reds_key,
            source=source,
        )

    # ── Transform ──────────────────────────────────────────────────────────────

    def transform(self) -> None:
        """Identify merge candidates (POSSIBLE_SAME_AS) before loading."""
        # Group persons by nome_normalized + data_nascimento
        by_nome_nasc: dict[str, list[REDSPerson]] = {}
        for p in self.persons.values():
            if not p.nome_normalized or not p.data_nascimento:
                continue
            key = f"{p.nome_normalized}|{p.data_nascimento}"
            by_nome_nasc.setdefault(key, []).append(p)

        # Flag persons with same name+nasc but different identity keys
        dups = 0
        for key, group in by_nome_nasc.items():
            if len(group) < 2:
                continue
            # Check if they have conflicting identity signals
            cpfs = {p.cpf for p in group if p.cpf}
            rgs = {p.rg_mg for p in group if p.rg_mg}
            if len(cpfs) > 1:
                logger.warning(
                    "[dhpp_reds] CONFLITO CPF: '%s' tem CPFs diferentes: %s",
                    key, cpfs
                )
            elif len(rgs) > 1:
                logger.warning(
                    "[dhpp_reds] CONFLITO RG: '%s' tem RGs diferentes: %s",
                    key, rgs
                )
            else:
                # Same name+nasc, no conflicting IDs → merge candidate
                self.merge_candidates.append({
                    "key": key,
                    "count": len(group),
                    "chaves": [p.chave_envolvido for p in group],
                })
                dups += 1

        logger.info(
            "[dhpp_reds] Transform complete. Merge candidates: %d", dups
        )

    # ── Load ───────────────────────────────────────────────────────────────────

    def load(self) -> None:
        loader = Neo4jBatchLoader(self.driver, batch_size=5_000)
        self._load_schema(loader)
        self._load_occurrences(loader)
        self._load_persons(loader)
        self._load_locations(loader)
        self._load_envolvido_rels(loader)
        self._load_occurrence_location_rels(loader)
        self._load_vehicle_nodes(loader)
        logger.info(
            "[dhpp_reds] Load complete. Written: %d rows total",
            loader._total_written,
        )
        if self.merge_candidates:
            logger.info(
                "[dhpp_reds] %d duplicate candidates queued for /intelink/links/review",
                len(self.merge_candidates),
            )

    def _load_schema(self, loader: Neo4jBatchLoader) -> None:
        """Apply constraints and indexes (idempotent)."""
        schema_file = (
            Path(__file__).parent.parent.parent.parent.parent.parent
            / "infra" / "neo4j" / "dhpp_reds_schema.cypher"
        )
        if not schema_file.exists():
            logger.warning("[dhpp_reds] Schema file not found: %s", schema_file)
            return
        cypher = schema_file.read_text()
        # Execute each statement separately (ignore IF NOT EXISTS errors)
        stmts = [s.strip() for s in cypher.split(";") if s.strip() and not s.strip().startswith("//")]
        with self.driver.session(database=self.neo4j_database) as session:
            for stmt in stmts:
                if stmt:
                    try:
                        session.run(stmt)
                    except Exception as e:
                        logger.debug("[dhpp_reds] Schema stmt skipped: %s", e)

    def _load_occurrences(self, loader: Neo4jBatchLoader) -> None:
        rows = [
            {
                "reds_number": o.reds_number,
                "type": o.type,
                "consumado": o.consumado,
                "data_fato": o.data_fato,
                "ano_fato": o.ano_fato,
                "mes_fato": o.mes_fato,
                "municipio": o.municipio,
                "bairro": o.bairro,
                "endereco": o.endereco,
                "latitude": o.latitude,
                "longitude": o.longitude,
                "modo_acao": o.modo_acao,
                "meio_utilizado": o.meio_utilizado,
                "causa_presumida": o.causa_presumida,
                "local_imediato": o.local_imediato,
                "historico": o.historico,
                "unidade_registro": o.unidade_registro,
                "orgao_registro": o.orgao_registro,
                "plates": o.plates,
                "calibers": o.calibers,
                "imeis": o.imeis,
                "source": o.source,
                "created_at": datetime.now(tz=UTC).isoformat(),
            }
            for o in self.occurrences.values()
        ]
        loader.run_query(
            """
            UNWIND $rows AS r
            MERGE (o:Occurrence {reds_number: r.reds_number})
            SET o += {
              type: r.type,
              consumado: r.consumado,
              data_fato: r.data_fato,
              ano_fato: r.ano_fato,
              mes_fato: r.mes_fato,
              municipio: r.municipio,
              bairro: r.bairro,
              endereco: r.endereco,
              latitude: r.latitude,
              longitude: r.longitude,
              modo_acao: r.modo_acao,
              meio_utilizado: r.meio_utilizado,
              causa_presumida: r.causa_presumida,
              local_imediato: r.local_imediato,
              historico: r.historico,
              unidade_registro: r.unidade_registro,
              orgao_registro: r.orgao_registro,
              plates: r.plates,
              calibers: r.calibers,
              imeis: r.imeis,
              source: r.source,
              created_at: r.created_at
            }
            """,
            rows,
        )
        logger.info("[dhpp_reds] Loaded %d :Occurrence nodes", len(rows))

    def _load_persons(self, loader: Neo4jBatchLoader) -> None:
        """Load :Person nodes with MG dedup strategy."""
        rows = []
        for p in self.persons.values():
            # Determine merge key priority: CPF > RG_MG > reds_person_key > chave_envolvido
            if p.cpf_valid and p.cpf:
                merge_key = "cpf"
                merge_val = p.cpf
            elif p.rg_mg:
                merge_key = "rg_mg"
                merge_val = p.rg_mg
            elif p.reds_person_key:
                merge_key = "reds_person_key"
                merge_val = p.reds_person_key
            else:
                merge_key = "chave_envolvido_key"
                merge_val = p.chave_envolvido

            rows.append({
                "merge_key": merge_key,
                "merge_val": merge_val,
                "chave_envolvido": p.chave_envolvido,
                "name": p.nome_normalized,
                "nome_original": p.nome,
                "sexo": p.sexo,
                "rg": p.rg,
                "rg_mg": p.rg_mg,
                "cpf": p.cpf,
                "nome_mae_normalized": p.nome_mae_normalized,
                "nome_mae": p.nome_mae,
                "data_nascimento": p.data_nascimento,
                "bairro": p.bairro,
                "municipio": p.municipio,
                "endereco": p.endereco,
                "telefone": p.telefone,
                "reds_person_key": p.reds_person_key,
                "source": p.source,
            })

        # Use CALL for each merge_key strategy
        # CPF-keyed persons
        cpf_rows = [r for r in rows if r["merge_key"] == "cpf"]
        if cpf_rows:
            loader.run_query(
                """
                UNWIND $rows AS r
                MERGE (p:Person {cpf: r.cpf})
                SET p += {
                  name: coalesce(p.name, r.name),
                  nome_original: r.nome_original,
                  sexo: coalesce(p.sexo, r.sexo),
                  rg: coalesce(p.rg, r.rg),
                  rg_mg: r.rg_mg,
                  nome_mae_normalized: coalesce(p.nome_mae_normalized, r.nome_mae_normalized),
                  nome_mae: coalesce(p.nome_mae, r.nome_mae),
                  data_nascimento: coalesce(p.data_nascimento, r.data_nascimento),
                  reds_person_key: r.reds_person_key,
                  source: r.source,
                  updated_at: datetime()
                }
                """,
                cpf_rows,
            )

        # RG-keyed persons (no CPF)
        rg_rows = [r for r in rows if r["merge_key"] == "rg_mg"]
        if rg_rows:
            loader.run_query(
                """
                UNWIND $rows AS r
                MERGE (p:Person {rg_mg: r.rg_mg})
                SET p += {
                  name: coalesce(p.name, r.name),
                  nome_original: r.nome_original,
                  sexo: coalesce(p.sexo, r.sexo),
                  rg: r.rg,
                  nome_mae_normalized: coalesce(p.nome_mae_normalized, r.nome_mae_normalized),
                  nome_mae: coalesce(p.nome_mae, r.nome_mae),
                  data_nascimento: coalesce(p.data_nascimento, r.data_nascimento),
                  reds_person_key: r.reds_person_key,
                  source: r.source,
                  updated_at: datetime()
                }
                """,
                rg_rows,
            )

        # reds_person_key persons (no CPF, no RG)
        key_rows = [r for r in rows if r["merge_key"] == "reds_person_key"]
        if key_rows:
            loader.run_query(
                """
                UNWIND $rows AS r
                MERGE (p:Person {reds_person_key: r.reds_person_key})
                SET p += {
                  name: coalesce(p.name, r.name),
                  nome_original: r.nome_original,
                  sexo: coalesce(p.sexo, r.sexo),
                  nome_mae_normalized: coalesce(p.nome_mae_normalized, r.nome_mae_normalized),
                  nome_mae: coalesce(p.nome_mae, r.nome_mae),
                  data_nascimento: coalesce(p.data_nascimento, r.data_nascimento),
                  source: r.source,
                  updated_at: datetime()
                }
                """,
                key_rows,
            )

        # Fallback: chave_envolvido (no identity signals at all)
        ck_rows = [r for r in rows if r["merge_key"] == "chave_envolvido_key"]
        if ck_rows:
            loader.run_query(
                """
                UNWIND $rows AS r
                MERGE (p:Person {chave_envolvido_key: r.chave_envolvido})
                SET p += {
                  name: coalesce(p.name, r.name),
                  nome_original: r.nome_original,
                  sexo: coalesce(p.sexo, r.sexo),
                  source: r.source,
                  updated_at: datetime()
                }
                """,
                ck_rows,
            )

        logger.info(
            "[dhpp_reds] Loaded %d :Person nodes (cpf=%d, rg=%d, key=%d, ck=%d)",
            len(rows), len(cpf_rows), len(rg_rows), len(key_rows), len(ck_rows),
        )

    def _load_locations(self, loader: Neo4jBatchLoader) -> None:
        """Upsert :Location nodes per bairro+municipio."""
        seen: set[str] = set()
        rows = []
        for o in self.occurrences.values():
            if not o.bairro:
                continue
            key = f"{o.bairro}|{o.municipio}"
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                "bairro": o.bairro,
                "municipio": o.municipio,
                "lat": o.latitude,
                "lon": o.longitude,
            })
        loader.run_query(
            """
            UNWIND $rows AS r
            MERGE (l:Location {bairro: r.bairro, municipio: r.municipio})
            SET l += {
              latitude: coalesce(l.latitude, r.lat),
              longitude: coalesce(l.longitude, r.lon),
              source: 'reds_patos_minas'
            }
            """,
            rows,
        )
        logger.info("[dhpp_reds] Loaded %d :Location nodes", len(rows))

    def _load_envolvido_rels(self, loader: Neo4jBatchLoader) -> None:
        """Create (Person)-[:ENVOLVIDO_EM]->(Occurrence) relationships."""
        rows = []
        for p in self.persons.values():
            # Determine which property to match on
            if p.cpf_valid and p.cpf:
                person_match = {"prop": "cpf", "val": p.cpf}
            elif p.rg_mg:
                person_match = {"prop": "rg_mg", "val": p.rg_mg}
            elif p.reds_person_key:
                person_match = {"prop": "reds_person_key", "val": p.reds_person_key}
            else:
                person_match = {"prop": "chave_envolvido_key", "val": p.chave_envolvido}

            rows.append({
                "person_prop": person_match["prop"],
                "person_val": person_match["val"],
                "reds_number": p.reds_number,
                "chave_envolvido": p.chave_envolvido,
                "tipo": p.tipo_envolvimento,
                "prisao": p.prisao,
                "grau_lesao": p.grau_lesao,
                "natureza_delito": p.natureza_delito,
            })

        # Run separately per prop to avoid dynamic property lookup issues
        for prop in ["cpf", "rg_mg", "reds_person_key", "chave_envolvido_key"]:
            batch = [r for r in rows if r["person_prop"] == prop]
            if not batch:
                continue
            loader.run_query(
                f"""
                UNWIND $rows AS r
                MATCH (p:Person {{{prop}: r.person_val}})
                MATCH (o:Occurrence {{reds_number: r.reds_number}})
                MERGE (p)-[rel:ENVOLVIDO_EM {{chave_envolvido: r.chave_envolvido}}]->(o)
                SET rel.tipo = r.tipo,
                    rel.prisao = r.prisao,
                    rel.grau_lesao = r.grau_lesao,
                    rel.natureza_delito = r.natureza_delito
                """,
                batch,
            )
        logger.info("[dhpp_reds] Loaded %d ENVOLVIDO_EM relationships", len(rows))

    def _load_occurrence_location_rels(self, loader: Neo4jBatchLoader) -> None:
        rows = [
            {"reds_number": o.reds_number, "bairro": o.bairro, "municipio": o.municipio}
            for o in self.occurrences.values()
            if o.bairro
        ]
        loader.run_query(
            """
            UNWIND $rows AS r
            MATCH (o:Occurrence {reds_number: r.reds_number})
            MATCH (l:Location {bairro: r.bairro, municipio: r.municipio})
            MERGE (o)-[:OCORREU_EM]->(l)
            """,
            rows,
        )
        logger.info("[dhpp_reds] Loaded %d OCORREU_EM relationships", len(rows))

    def _load_vehicle_nodes(self, loader: Neo4jBatchLoader) -> None:
        """Create :Vehicle nodes for plates extracted from histórico."""
        rows = []
        for o in self.occurrences.values():
            for plate in o.plates:
                rows.append({
                    "plate": plate,
                    "reds_number": o.reds_number,
                    "source": o.source,
                })
        if not rows:
            return
        loader.run_query(
            """
            UNWIND $rows AS r
            MERGE (v:Vehicle {plate: r.plate})
            SET v.source = coalesce(v.source, r.source)
            WITH v, r
            MATCH (o:Occurrence {reds_number: r.reds_number})
            MERGE (v)-[:MENCIONADO_EM]->(o)
            """,
            rows,
        )
        logger.info("[dhpp_reds] Loaded %d :Vehicle nodes from plates", len(rows))


# ── Utility helpers ────────────────────────────────────────────────────────────

def _safe_float(val: Any) -> float | None:
    try:
        return float(val) if val is not None else None
    except (ValueError, TypeError):
        return None


def _normalize_prisao(raw: str) -> str | None:
    if not raw:
        return None
    if "FLAGRANTE" in raw:
        return "FLAGRANTE"
    if "SEM PRISAO" in raw or "SEM PRISÃO" in raw:
        return "SEM_PRISAO"
    if "APREEND" in raw:
        return "APREENDIDO"
    if "MANDADO" in raw:
        return "MANDADO"
    return raw[:50] if raw else None


def _best_phone(*phones: Any) -> str | None:
    for p in phones:
        s = str(p or "").strip()
        if s and s not in ("-", "None", ""):
            return s
    return None
